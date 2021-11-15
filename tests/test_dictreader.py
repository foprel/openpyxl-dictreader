""" unit tests for openpyxl DictReader
    heavily leveraged from the CPython tests """

import unittest
from tempfile import TemporaryFile
from itertools import permutations
from textwrap import dedent
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl_dictreader import DictReader


class TestDictFields(unittest.TestCase):
    ### "long" means the row is longer than the number of fieldnames
    ### "short" means there are fewer elements in the row than fieldnames

    def test_read_dict_fields(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2", "f3"])
            self.assertEqual(next(reader), {"f1": "1", "f2": "2", "f3": "abc"})

    def test_read_dict_no_fieldnames(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["C1"] = "f3"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            self.assertEqual(next(reader), {"f1": "1", "f2": "2", "f3": "abc"})
            self.assertEqual(reader.fieldnames, ["f1", "f2", "f3"])

    def test_read_dict_fieldnames_chain(self):
        import itertools

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["C1"] = "f3"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            first = next(reader)
            for row in itertools.chain([first], reader):
                self.assertEqual(reader.fieldnames, ["f1", "f2", "f3"])
                self.assertEqual(row, {"f1": "1", "f2": "2", "f3": "abc"})

    def test_read_long(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2"])
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", None: ["abc", "4", "5", "6"]}
            )

    def test_read_long_with_rest(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2"], restkey="_rest")
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", "_rest": ["abc", "4", "5", "6"]}
            )

    # The dict reader does not deal with 'restkey' yet.
    #    It is expecting blank return values but is getting 'None'

    #  E           AssertionError: Lists differ: ['f1', 'f2', None, None, None, None] != ['f1', 'f2']
    #  E           First list contains 4 additional elements.

    @unittest.skip("DictReader does not handle 'restkey' option yet")
    def test_read_long_with_rest_no_fieldnames(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        ws["D2"] = "4"
        ws["E2"] = "5"
        ws["F2"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, restkey="_rest")
            self.assertEqual(reader.fieldnames, ["f1", "f2"])
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", "_rest": ["abc", "4", "5", "6"]}
            )

    # The dict reader does not deal with 'restval' yet.
    #    It is expecting blank return values but is getting 'None'
    @unittest.skip("DictReader does not handle 'restval' option yet")
    def test_read_short(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(
                fileobj, fieldnames="1 2 3 4 5 6".split(), restval="DEFAULT"
            )
            self.assertEqual(
                next(reader),
                {"1": "1", "2": "2", "3": "abc", "4": "4", "5": "5", "6": "6"},
            )
            self.assertEqual(
                next(reader),
                {
                    "1": "1",
                    "2": "2",
                    "3": "abc",
                    "4": "DEFAULT",
                    "5": "DEFAULT",
                    "6": "DEFAULT",
                },
            )

    def test_read_multi(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "2147483648"
        ws["B1"] = "43.0e12"
        ws["C1"] = "17"
        ws["D1"] = "abc"
        ws["E1"] = "def"
        ws["A2"] = "147483648"
        ws["B2"] = "43.0e12"
        ws["C2"] = "17"
        ws["D2"] = "abc"
        ws["E2"] = "def"
        ws["A3"] = "47483648"
        ws["B3"] = "43.0"
        ws["C3"] = "170"
        ws["D3"] = "abc"
        ws["E3"] = "def"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames="i1 float i2 s1 s2".split())
            self.assertEqual(
                next(reader),
                {
                    "i1": "2147483648",
                    "float": "43.0e12",
                    "i2": "17",
                    "s1": "abc",
                    "s2": "def",
                },
            )

    @unittest.skip("DictReader does not handle blank lines yet")
    def test_read_with_blanks(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"

        ws["A3"] = "1"
        ws["B3"] = "2"
        ws["C3"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(
                fileobj,
                fieldnames="1 2 3".split(),
            )
            self.assertEqual(next(reader), {"1": "1", "2": "2", "3": "abc"})
            self.assertEqual(next(reader), {"1": "1", "2": "2", "3": "abc"})


class KeyOrderingTest(unittest.TestCase):
    @unittest.skip("requires DictWriter, which is not ready yet")
    def test_ordering_for_the_dict_reader_and_writer(self):
        resultset = set()
        for keys in permutations("abcde"):
            with TemporaryFile("w+", newline="", encoding="utf-8") as fileobject:
                dw = csv.DictWriter(fileobject, keys)
                dw.writeheader()
                fileobject.seek(0)
                dr = csv.DictReader(fileobject)
                kt = tuple(dr.fieldnames)
                self.assertEqual(keys, kt)
                resultset.add(kt)
        # Final sanity check: were all permutations unique?
        self.assertEqual(
            len(resultset),
            120,
            "Key ordering: some key permutations not collected (expected 120)",
        )

    @unittest.skip("OrderedDict no longer used")
    def test_ordered_dict_reader(self):
        data = dedent(
            """\
            FirstName,LastName
            Eric,Idle
            Graham,Chapman,Over1,Over2
            Under1
            John,Cleese
        """
        ).splitlines()

        self.assertEqual(
            list(csv.DictReader(data)),
            [
                OrderedDict([("FirstName", "Eric"), ("LastName", "Idle")]),
                OrderedDict(
                    [
                        ("FirstName", "Graham"),
                        ("LastName", "Chapman"),
                        (None, ["Over1", "Over2"]),
                    ]
                ),
                OrderedDict([("FirstName", "Under1"), ("LastName", None)]),
                OrderedDict([("FirstName", "John"), ("LastName", "Cleese")]),
            ],
        )

        self.assertEqual(
            list(csv.DictReader(data, restkey="OtherInfo")),
            [
                OrderedDict([("FirstName", "Eric"), ("LastName", "Idle")]),
                OrderedDict(
                    [
                        ("FirstName", "Graham"),
                        ("LastName", "Chapman"),
                        ("OtherInfo", ["Over1", "Over2"]),
                    ]
                ),
                OrderedDict([("FirstName", "Under1"), ("LastName", None)]),
                OrderedDict([("FirstName", "John"), ("LastName", "Cleese")]),
            ],
        )

        del data[0]  # Remove the header row
        self.assertEqual(
            list(csv.DictReader(data, fieldnames=["fname", "lname"])),
            [
                OrderedDict([("fname", "Eric"), ("lname", "Idle")]),
                OrderedDict(
                    [
                        ("fname", "Graham"),
                        ("lname", "Chapman"),
                        (None, ["Over1", "Over2"]),
                    ]
                ),
                OrderedDict([("fname", "Under1"), ("lname", None)]),
                OrderedDict([("fname", "John"), ("lname", "Cleese")]),
            ],
        )
