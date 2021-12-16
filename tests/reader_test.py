""" unit tests for openpyxl DictReader
    heavily leveraged from the CPython tests """

import unittest
from io import BytesIO
from tempfile import TemporaryFile
from openpyxl import Workbook
from openpyxl_dictreader import DictReader


class TestDictFields(unittest.TestCase):
    ### "long" means the row is longer than the number of fieldnames
    ### "short" means there are fewer elements in the row than fieldnames

    def test_read_dict_fields(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2", "f3"])
            self.assertEqual(next(reader), {"f1": "1", "f2": "2", "f3": "abc"})

    def test_read_dict_no_fieldnames(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["C1"] = "f3"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            self.assertEqual(next(reader), {"f1": "1", "f2": "2", "f3": "abc"})
            self.assertEqual(reader.fieldnames, ["f1", "f2", "f3"])

    def test_read_dict_set_fieldnames(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["B1"] = 2
        ws["C1"] = 3
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            reader.fieldnames = ['John', 'Sebastian', 'Bach']
            self.assertEqual(next(reader), {"John": 1, "Sebastian": 2, "Bach": 3})
            self.assertEqual(reader.fieldnames, ['John', 'Sebastian', 'Bach'])

    def test_read_dict_fieldnames_from_blank_file(self):
        wb = Workbook()
        ws = wb.active
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            self.assertIsNone(reader.fieldnames)

    def test_read_dict_fieldnames_chain(self):
        import itertools

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["C1"] = "f3"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            first = next(reader)
            for row in itertools.chain([first], reader):
                self.assertEqual(reader.fieldnames, ["f1", "f2", "f3"])
                self.assertEqual(row, {"f1": "1", "f2": "2", "f3": "abc"})

    def test_read_long(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2"])
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", None: ["abc", "4", "5", "6"]}
            )

    def test_read_long_with_rest(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames=["f1", "f2"], restkey="_rest")
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", "_rest": ["abc", "4", "5", "6"]}
            )

    def test_read_long_with_rest_no_fieldnames(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "f1"
        ws["B1"] = "f2"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        ws["D2"] = "4"
        ws["E2"] = "5"
        ws["F2"] = "6"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, restkey="_rest")
            self.assertEqual(reader.fieldnames, ["f1", "f2"])
            self.assertEqual(
                next(reader), {"f1": "1", "f2": "2", "_rest": ["abc", "4", "5", "6"]}
            )

    def test_read_short(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"
        ws["D1"] = "4"
        ws["E1"] = "5"
        ws["F1"] = "6"
        ws["A2"] = "1"
        ws["B2"] = "2"
        ws["C2"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(
                fileobj, fieldnames="1 2 3 4 5 6".split(), restval="DEFAULT"
            )
            self.assertEqual(
                next(reader),
                {"1": "1", "2": "2", "3": "abc", "4": "4", "5": "5", "6": "6"},
            )
            self.assertEqual(
                next(reader),
                {
                    "1": "1",
                    "2": "2",
                    "3": "abc",
                    "4": "DEFAULT",
                    "5": "DEFAULT",
                    "6": "DEFAULT",
                },
            )

    def test_read_multi(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "2147483648"
        ws["B1"] = "43.0e12"
        ws["C1"] = "17"
        ws["D1"] = "abc"
        ws["E1"] = "def"
        ws["A2"] = "147483648"
        ws["B2"] = "43.0e12"
        ws["C2"] = "17"
        ws["D2"] = "abc"
        ws["E2"] = "def"
        ws["A3"] = "47483648"
        ws["B3"] = "43.0"
        ws["C3"] = "170"
        ws["D3"] = "abc"
        ws["E3"] = "def"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj, fieldnames="i1 float i2 s1 s2".split())
            self.assertEqual(
                next(reader),
                {
                    "i1": "2147483648",
                    "float": "43.0e12",
                    "i2": "17",
                    "s1": "abc",
                    "s2": "def",
                },
            )

    def test_read_with_blanks(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "1"
        ws["B1"] = "2"
        ws["C1"] = "abc"

        ws["A3"] = "1"
        ws["B3"] = "2"
        ws["C3"] = "abc"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(
                fileobj,
                fieldnames="1 2 3".split(),
            )
            self.assertEqual(next(reader), {"1": "1", "2": "2", "3": "abc"})
            self.assertEqual(next(reader), {"1": "1", "2": "2", "3": "abc"})

    def test_read_with_blank_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "veg"
        ws["B1"] = "fruit"
        ws["C1"] = "meat"

        ws["A2"] = "carrot"
        ws["C2"] = "mutton"
        with TemporaryFile() as fileobj:
            wb.save(fileobj)
            reader = DictReader(fileobj)
            self.assertEqual(next(reader), {"veg": "carrot", "fruit": None, "meat": "mutton"})


class KeyOrderingTest(unittest.TestCase):

    def test_ordered_dict_reader(self):
        stream = BytesIO()
        rows = (
            ("FirstName", "LastName"),
            ("Eric", "Idle"),
            ("Graham", "Chapman", "Over1", "Over2"),
            (),
            ("Under1",),
            ("John", "Cleese"),
        )
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(stream)

        self.assertEqual(
            list(DictReader(stream)),
            [
                dict([("FirstName", "Eric"), ("LastName", "Idle")]),
                dict(
                    [
                        ("FirstName", "Graham"),
                        ("LastName", "Chapman"),
                        (None, ["Over1", "Over2"]),
                    ]
                ),
                dict([("FirstName", "Under1"), ("LastName", None)]),
                dict([("FirstName", "John"), ("LastName", "Cleese")]),
            ],
        )

        self.assertEqual(
            list(DictReader(stream, restkey="OtherInfo")),
            [
                dict([("FirstName", "Eric"), ("LastName", "Idle")]),
                dict(
                    [
                        ("FirstName", "Graham"),
                        ("LastName", "Chapman"),
                        ("OtherInfo", ["Over1", "Over2"]),
                    ]
                ),
                dict([("FirstName", "Under1"), ("LastName", None)]),
                dict([("FirstName", "John"), ("LastName", "Cleese")]),
            ],
        )

        # test with header-less file
        stream = BytesIO()
        wb = Workbook()
        ws = wb.active
        for row in rows[1:]: # skip the header
            ws.append(row)
        wb.save(stream)
        self.assertEqual(
            list(DictReader(stream, fieldnames=["fname", "lname"])),
            [
                dict([("fname", "Eric"), ("lname", "Idle")]),
                dict(
                    [
                        ("fname", "Graham"),
                        ("lname", "Chapman"),
                        (None, ["Over1", "Over2"]),
                    ]
                ),
                dict([("fname", "Under1"), ("lname", None)]),
                dict([("fname", "John"), ("lname", "Cleese")]),
            ],
        )

